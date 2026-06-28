# Registration addon routes - add to your existing FastAPI app with:
# from registration_routes import router as registration_router
# app.include_router(registration_router)

import json
import os
import re
import shutil
import uuid
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Form, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from starlette.datastructures import UploadFile as StarletteUploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("REGISTRATION_DATA_DIR", APP_DIR / "registration_data"))
UPLOAD_DIR = DATA_DIR / "uploads"
DATA_FILE = DATA_DIR / "registrations.json"
WHATSAPP_GROUPS_FILE = DATA_DIR / "whatsapp_groups.json"
MAX_TEAMS = int(os.getenv("REGISTRATION_MAX_TEAMS", "12"))
ADMIN_PASSWORD = os.getenv("REGISTRATION_ADMIN_PASSWORD", "BeshooWarZone")
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "8"))
MAX_UPLOAD_BYTES = MAX_UPLOAD_MB * 1024 * 1024

DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

router = APIRouter()


# ----- HTML pages for the registration addon -----
@router.get("/register", include_in_schema=False)
def serve_register_page():
    page = APP_DIR / "register.html"
    if not page.exists():
        raise HTTPException(status_code=404, detail="register.html غير موجود بجانب main.py")
    return FileResponse(page)


@router.get("/registrations", include_in_schema=False)
def serve_registrations_page():
    page = APP_DIR / "registrations.html"
    if not page.exists():
        raise HTTPException(status_code=404, detail="registrations.html غير موجود بجانب main.py")
    return FileResponse(page)


@router.get("/whatsapp-groups", include_in_schema=False)
def serve_whatsapp_groups_page():
    page = APP_DIR / "whatsapp_groups.html"
    if not page.exists():
        raise HTTPException(status_code=404, detail="whatsapp_groups.html غير موجود بجانب main.py")
    return FileResponse(page)


@router.get("/register.html", include_in_schema=False)
def serve_register_html_alias():
    return serve_register_page()


@router.get("/registrations.html", include_in_schema=False)
def serve_registrations_html_alias():
    return serve_registrations_page()


@router.get("/whatsapp-groups.html", include_in_schema=False)
def serve_whatsapp_groups_html_alias():
    return serve_whatsapp_groups_page()


ALLOWED_IMAGE_TYPES = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp"}
REQUIRED_PLAYER_FIELDS = ["name", "age", "birthdate", "national_id", "university", "college", "gender"]
FILE_FIELDS = ["photo", "id_card", "university_card"]

ARABIC_DIGIT_TRANSLATION = str.maketrans({
    "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
    "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9",
    "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4",
    "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
})


def to_english_digits(value: Any) -> str:
    """Accept Arabic/English digits and return a string with English digits only."""
    if value is None:
        return ""
    return str(value).translate(ARABIC_DIGIT_TRANSLATION).strip()


def normalize_birthdate(value: Any) -> str:
    raw = to_english_digits(value)
    raw = raw.replace("/", "-").replace(".", "-").replace("–", "-").replace("—", "-")
    raw = re.sub(r"\s+", "", raw)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    raise ValueError("invalid date")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_data() -> Dict[str, Any]:
    if not DATA_FILE.exists():
        return {"teams": []}
    try:
        with DATA_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if "teams" not in data or not isinstance(data["teams"], list):
            return {"teams": []}
        return data
    except Exception:
        return {"teams": []}


def save_data(data: Dict[str, Any]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    tmp = DATA_FILE.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp.replace(DATA_FILE)


def default_whatsapp_groups() -> Dict[str, Any]:
    return {
        "groups": [
            {"slot": i, "name": f"جروب واتساب {i}", "link": "", "team_id": None}
            for i in range(1, MAX_TEAMS + 1)
        ]
    }


def load_whatsapp_groups() -> Dict[str, Any]:
    if not WHATSAPP_GROUPS_FILE.exists():
        data = default_whatsapp_groups()
        save_whatsapp_groups(data)
        return data
    try:
        with WHATSAPP_GROUPS_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = default_whatsapp_groups()

    groups = data.get("groups") if isinstance(data, dict) else None
    if not isinstance(groups, list):
        groups = []

    by_slot = {}
    for g in groups:
        try:
            slot = int(g.get("slot"))
        except Exception:
            continue
        if 1 <= slot <= MAX_TEAMS:
            by_slot[slot] = {
                "slot": slot,
                "name": str(g.get("name") or f"جروب واتساب {slot}"),
                "link": str(g.get("link") or "").strip(),
                "team_id": g.get("team_id") or None,
            }

    fixed = []
    for i in range(1, MAX_TEAMS + 1):
        fixed.append(by_slot.get(i) or {"slot": i, "name": f"جروب واتساب {i}", "link": "", "team_id": None})
    return {"groups": fixed}


def save_whatsapp_groups(data: Dict[str, Any]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    tmp = WHATSAPP_GROUPS_FILE.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp.replace(WHATSAPP_GROUPS_FILE)


def group_for_team(team_id: str) -> Optional[Dict[str, Any]]:
    for group in load_whatsapp_groups().get("groups", []):
        if group.get("team_id") == team_id:
            return group
    return None


def release_whatsapp_group(team_id: str) -> None:
    groups_data = load_whatsapp_groups()
    changed = False
    for group in groups_data.get("groups", []):
        if group.get("team_id") == team_id:
            group["team_id"] = None
            changed = True
    if changed:
        save_whatsapp_groups(groups_data)


def assign_whatsapp_group(team_id: str) -> Dict[str, Any]:
    groups_data = load_whatsapp_groups()

    # لو الفريق كان متسجل قبل كده، رجّع نفس الجروب.
    for group in groups_data.get("groups", []):
        if group.get("team_id") == team_id:
            return group

    # لازم يكون في لينك محفوظ وغير مت assigned.
    for group in groups_data.get("groups", []):
        if not group.get("team_id") and str(group.get("link") or "").strip():
            group["team_id"] = team_id
            save_whatsapp_groups(groups_data)
            return group

    raise HTTPException(status_code=400, detail="لا يوجد لينك جروب واتساب متاح. ادخل 12 لينك من صفحة إدارة جروبات الواتساب أولًا.")


def normalize_team_name(name: str) -> str:
    name = re.sub(r"\s+", " ", (name or "").strip())
    return name.casefold()


def slugify(value: str) -> str:
    value = re.sub(r"[^\w\-]+", "_", value.strip(), flags=re.UNICODE)
    return value[:80] or "file"


def require_admin(request: Request) -> None:
    supplied = request.headers.get("x-admin-password") or request.query_params.get("p") or request.query_params.get("password")
    if supplied != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")


def public_team(team: Dict[str, Any], request: Optional[Request] = None, include_files: bool = True) -> Dict[str, Any]:
    item = json.loads(json.dumps(team, ensure_ascii=False))
    if include_files and request is not None:
        for player in item.get("players", []):
            files = player.get("files", {})
            player["file_urls"] = {}
            for kind in FILE_FIELDS:
                if files.get(kind):
                    player["file_urls"][kind] = str(request.base_url).rstrip("/") + f"/api/registration-file/{item['id']}/{player['id']}/{kind}"
    return item


def validate_players(players: List[Dict[str, Any]], old_team: Optional[Dict[str, Any]] = None) -> None:
    if not (12 <= len(players) <= 15):
        raise HTTPException(status_code=400, detail="كل فريق لازم يكون من 12 إلى 15 فرد.")

    male_count = 0
    female_count = 0
    seen_national_ids = set()

    for idx, player in enumerate(players, start=1):
        for field in REQUIRED_PLAYER_FIELDS:
            if not str(player.get(field, "")).strip():
                raise HTTPException(status_code=400, detail=f"بيانات اللاعب رقم {idx} ناقصة: {field}")

        try:
            age_text = re.sub(r"\D+", "", to_english_digits(player.get("age")))
            age = int(age_text)
            if age <= 0 or age > 100:
                raise ValueError
            player["age"] = age
        except Exception:
            raise HTTPException(status_code=400, detail=f"سن اللاعب رقم {idx} غير صحيح.")

        national_id = re.sub(r"\D+", "", to_english_digits(player.get("national_id", "")))
        if len(national_id) != 14:
            raise HTTPException(status_code=400, detail=f"الرقم القومي للاعب رقم {idx} لازم يكون 14 رقم.")
        if national_id in seen_national_ids:
            raise HTTPException(status_code=400, detail=f"الرقم القومي مكرر داخل نفس الفريق عند اللاعب رقم {idx}.")
        seen_national_ids.add(national_id)
        player["national_id"] = national_id

        gender = str(player.get("gender", "")).strip().lower()
        if gender in ["male", "ذكر", "m"]:
            player["gender"] = "ذكر"
            male_count += 1
        elif gender in ["female", "انثى", "أنثى", "f"]:
            player["gender"] = "أنثى"
            female_count += 1
        else:
            raise HTTPException(status_code=400, detail=f"نوع اللاعب رقم {idx} لازم يكون ذكر أو أنثى.")

        # Basic date format check + normalize all date digits to English
        try:
            player["birthdate"] = normalize_birthdate(player.get("birthdate"))
        except Exception:
            raise HTTPException(status_code=400, detail=f"تاريخ ميلاد اللاعب رقم {idx} غير صحيح.")

    if male_count < 6 or female_count < 6:
        raise HTTPException(status_code=400, detail="كل فريق لازم يكون فيه على الأقل 6 ذكور و6 إناث.")


def ensure_team_name_unique(data: Dict[str, Any], team_name: str, exclude_team_id: Optional[str] = None) -> None:
    normalized = normalize_team_name(team_name)
    if not normalized:
        raise HTTPException(status_code=400, detail="اسم المنتخب مطلوب.")
    for team in data.get("teams", []):
        if exclude_team_id and team.get("id") == exclude_team_id:
            continue
        if normalize_team_name(team.get("team_name", "")) == normalized:
            raise HTTPException(status_code=409, detail="اسم المنتخب مستخدم قبل كده، اختار اسم تاني.")


async def save_uploaded_file(upload: StarletteUploadFile, team_id: str, player_id: str, kind: str) -> str:
    if not upload or not getattr(upload, "filename", ""):
        raise HTTPException(status_code=400, detail=f"ملف {kind} مطلوب.")

    content_type = upload.content_type or ""
    ext = ALLOWED_IMAGE_TYPES.get(content_type)
    if not ext:
        raise HTTPException(status_code=400, detail="الصور المسموحة: JPG / PNG / WEBP فقط.")

    data = await upload.read(MAX_UPLOAD_BYTES + 1)
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail=f"حجم الصورة لا يزيد عن {MAX_UPLOAD_MB}MB.")

    folder = UPLOAD_DIR / team_id / player_id
    folder.mkdir(parents=True, exist_ok=True)
    filename = f"{kind}_{uuid.uuid4().hex}{ext}"
    path = folder / filename
    with path.open("wb") as f:
        f.write(data)
    return str(path.relative_to(DATA_DIR))


def delete_team_files(team_id: str) -> None:
    folder = UPLOAD_DIR / team_id
    if folder.exists():
        shutil.rmtree(folder, ignore_errors=True)


async def build_team_from_form(request: Request, existing: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    form = await request.form()
    team_name = str(form.get("team_name", "")).strip()
    players_json = str(form.get("players_json", "[]"))

    try:
        players_raw = json.loads(players_json)
    except Exception:
        raise HTTPException(status_code=400, detail="صيغة بيانات اللاعبين غير صحيحة.")

    if not isinstance(players_raw, list):
        raise HTTPException(status_code=400, detail="بيانات اللاعبين لازم تكون قائمة.")

    team_id = existing.get("id") if existing else uuid.uuid4().hex
    old_players_by_id = {p.get("id"): p for p in (existing or {}).get("players", [])}
    players: List[Dict[str, Any]] = []

    for idx, raw in enumerate(players_raw):
        player_id = raw.get("id") or uuid.uuid4().hex
        client_key = str(raw.get("client_key") or raw.get("id") or idx)
        old_player = old_players_by_id.get(player_id, {})
        old_files = old_player.get("files", {})

        player = {
            "id": player_id,
            "name": str(raw.get("name", "")).strip(),
            "age": to_english_digits(raw.get("age", "")),
            "birthdate": to_english_digits(raw.get("birthdate", "")),
            "national_id": to_english_digits(raw.get("national_id", "")),
            "university": str(raw.get("university", "")).strip(),
            "college": str(raw.get("college", "")).strip(),
            "gender": str(raw.get("gender", "")).strip(),
            "files": dict(old_files),
        }

        for kind in FILE_FIELDS:
            upload = form.get(f"{kind}_{client_key}")
            if isinstance(upload, StarletteUploadFile) and upload.filename:
                player["files"][kind] = await save_uploaded_file(upload, team_id, player_id, kind)
            elif not player["files"].get(kind):
                raise HTTPException(status_code=400, detail=f"صورة {kind} مطلوبة للاعب رقم {idx + 1}.")

        players.append(player)

    validate_players(players, old_team=existing)

    return {
        "id": team_id,
        "team_name": team_name,
        "created_at": (existing or {}).get("created_at") or now_iso(),
        "updated_at": now_iso(),
        "players": players,
    }



@router.get("/api/whatsapp-groups")
def get_whatsapp_groups(request: Request):
    require_admin(request)
    data = load_data()
    teams_by_id = {t.get("id"): t.get("team_name") for t in data.get("teams", [])}
    groups = []
    for g in load_whatsapp_groups().get("groups", []):
        item = dict(g)
        item["assigned_team_name"] = teams_by_id.get(g.get("team_id"))
        groups.append(item)
    return {"groups": groups, "max_teams": MAX_TEAMS}


@router.post("/api/whatsapp-groups")
async def save_whatsapp_groups_api(request: Request):
    require_admin(request)
    payload = await request.json()
    incoming = payload.get("groups") if isinstance(payload, dict) else None
    if not isinstance(incoming, list):
        raise HTTPException(status_code=400, detail="صيغة لينكات الواتساب غير صحيحة.")

    existing = load_whatsapp_groups()
    existing_by_slot = {int(g.get("slot")): g for g in existing.get("groups", [])}
    final = []
    for i in range(1, MAX_TEAMS + 1):
        old = existing_by_slot.get(i, {"slot": i, "name": f"جروب واتساب {i}", "link": "", "team_id": None})
        sent = next((g for g in incoming if str(g.get("slot")) == str(i)), {})
        link = str(sent.get("link", old.get("link", ""))).strip()
        name = str(sent.get("name", old.get("name") or f"جروب واتساب {i}")).strip() or f"جروب واتساب {i}"
        if link and not re.match(r"^https?://", link):
            raise HTTPException(status_code=400, detail=f"لينك الجروب رقم {i} لازم يبدأ بـ http أو https.")
        final.append({"slot": i, "name": name, "link": link, "team_id": old.get("team_id") or None})

    save_whatsapp_groups({"groups": final})
    return {"status": "success", "groups": final}


@router.get("/api/team-name-available")
def team_name_available(name: str):
    data = load_data()
    normalized = normalize_team_name(name)
    if not normalized:
        return {"available": False, "message": "اكتب اسم المنتخب"}
    for team in data.get("teams", []):
        if normalize_team_name(team.get("team_name", "")) == normalized:
            return {"available": False, "message": "الاسم مستخدم قبل كده"}
    return {"available": True, "message": "الاسم متاح"}


@router.post("/api/register-team")
async def register_team(request: Request):
    data = load_data()
    if len(data.get("teams", [])) >= MAX_TEAMS:
        raise HTTPException(status_code=400, detail=f"تم اكتمال عدد الفرق المسموح به: {MAX_TEAMS} فريق.")

    team = await build_team_from_form(request)
    ensure_team_name_unique(data, team["team_name"])

    # Check national ID uniqueness across teams.
    existing_ids = {p.get("national_id") for t in data.get("teams", []) for p in t.get("players", [])}
    for p in team["players"]:
        if p.get("national_id") in existing_ids:
            delete_team_files(team["id"])
            raise HTTPException(status_code=409, detail=f"الرقم القومي {p.get('national_id')} مسجل قبل كده في فريق آخر.")

    try:
        whatsapp_group = assign_whatsapp_group(team["id"])
    except Exception:
        delete_team_files(team["id"])
        raise

    team["whatsapp_group_slot"] = whatsapp_group.get("slot")
    data["teams"].append(team)
    save_data(data)
    return {
        "status": "success",
        "team_id": team["id"],
        "team_name": team["team_name"],
        "whatsapp_group": {
            "slot": whatsapp_group.get("slot"),
            "name": whatsapp_group.get("name"),
            "link": whatsapp_group.get("link"),
        },
    }


@router.get("/api/registrations")
def list_registrations(request: Request):
    require_admin(request)
    data = load_data()
    groups_by_team = {g.get("team_id"): g for g in load_whatsapp_groups().get("groups", []) if g.get("team_id")}
    teams = []
    for team in data.get("teams", []):
        males = sum(1 for p in team.get("players", []) if p.get("gender") == "ذكر")
        females = sum(1 for p in team.get("players", []) if p.get("gender") == "أنثى")
        whatsapp = groups_by_team.get(team.get("id"))
        teams.append({
            "id": team.get("id"),
            "team_name": team.get("team_name"),
            "created_at": team.get("created_at"),
            "updated_at": team.get("updated_at"),
            "players_count": len(team.get("players", [])),
            "males": males,
            "females": females,
            "whatsapp_group": {
                "slot": whatsapp.get("slot"),
                "name": whatsapp.get("name"),
                "link": whatsapp.get("link"),
            } if whatsapp else None,
        })
    return {"teams": teams, "max_teams": MAX_TEAMS}


@router.get("/api/registrations/export")
def export_registrations(request: Request):
    require_admin(request)
    data = load_data()
    wb = Workbook()

    # المطلوب: كل فريق في Sheet لوحده، وبس الأعمدة الأساسية:
    # الاسم، السن، تاريخ الميلاد.
    default_ws = wb.active
    wb.remove(default_ws)

    def safe_sheet_title(title: str, used: set) -> str:
        title = str(title or "Team").strip() or "Team"
        # Excel لا يسمح بهذه الرموز في اسم الشيت: : \/ ? * [ ]
        title = re.sub(r"[:\\/\?\*\[\]]+", "-", title)
        title = re.sub(r"\s+", " ", title).strip()
        title = title[:31] or "Team"
        base = title
        counter = 2
        while title in used:
            suffix = f" {counter}"
            title = (base[:31-len(suffix)] + suffix)[:31]
            counter += 1
        used.add(title)
        return title

    used_titles = set()
    teams = data.get("teams", [])

    if not teams:
        ws = wb.create_sheet("No registrations")
        ws.append(["اسم اللاعب", "السن", "تاريخ الميلاد"])
    else:
        for team in teams:
            ws = wb.create_sheet(safe_sheet_title(team.get("team_name", "Team"), used_titles))
            ws.append(["اسم اللاعب", "السن", "تاريخ الميلاد"])
            for player in team.get("players", []):
                birthdate = player.get("birthdate", "")
                try:
                    birthdate = normalize_birthdate(birthdate)
                except Exception:
                    birthdate = to_english_digits(birthdate)
                ws.append([
                    player.get("name", ""),
                    to_english_digits(player.get("age", "")),
                    birthdate,
                ])

            # تنسيق بسيط لكل شيت.
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1E3A8A")
                cell.alignment = Alignment(horizontal="center")

            widths = {"A": 28, "B": 12, "C": 18}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"warzone_registrations_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/api/registrations/{team_id}/download")
def download_team_package(team_id: str, request: Request):
    """Download one team's data and all uploaded images as a ZIP file."""
    require_admin(request)
    data = load_data()
    team = next((t for t in data.get("teams", []) if t.get("id") == team_id), None)
    if not team:
        raise HTTPException(status_code=404, detail="الفريق غير موجود.")

    def clean_name(value: str, fallback: str = "item") -> str:
        value = str(value or fallback).strip()
        value = re.sub(r"[\\/:*?\"<>|]+", "-", value)
        value = re.sub(r"\s+", " ", value).strip()
        return value[:90] or fallback

    # Excel file inside the ZIP contains the full team data.
    wb = Workbook()
    ws = wb.active
    ws.title = "Team Data"
    headers = [
        "رقم", "اسم اللاعب", "السن", "تاريخ الميلاد", "النوع",
        "الرقم القومي", "الجامعة", "الكلية",
        "الصورة الشخصية", "صورة البطاقة", "صورة كارنيه الجامعة",
    ]
    ws.append(headers)

    file_labels = {
        "photo": "الصورة الشخصية",
        "id_card": "صورة البطاقة",
        "university_card": "صورة كارنيه الجامعة",
    }

    for idx, player in enumerate(team.get("players", []), start=1):
        files = player.get("files", {}) or {}
        ws.append([
            idx,
            player.get("name", ""),
            to_english_digits(player.get("age", "")),
            to_english_digits(player.get("birthdate", "")),
            player.get("gender", ""),
            to_english_digits(player.get("national_id", "")),
            player.get("university", ""),
            player.get("college", ""),
            Path(files.get("photo", "")).name if files.get("photo") else "",
            Path(files.get("id_card", "")).name if files.get("id_card") else "",
            Path(files.get("university_card", "")).name if files.get("university_card") else "",
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")
        cell.alignment = Alignment(horizontal="center")
    for col, width in {
        "A": 8, "B": 28, "C": 10, "D": 18, "E": 12,
        "F": 20, "G": 22, "H": 22, "I": 28, "J": 28, "K": 32,
    }.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    xlsx_bytes = BytesIO()
    wb.save(xlsx_bytes)
    xlsx_bytes.seek(0)

    zip_buffer = BytesIO()
    team_folder = clean_name(team.get("team_name") or team_id, "team")
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # Save full JSON backup too, in case you need to restore/edit manually.
        zf.writestr(f"{team_folder}/team_data.json", json.dumps(team, ensure_ascii=False, indent=2))
        zf.writestr(f"{team_folder}/team_data.xlsx", xlsx_bytes.getvalue())

        # Add uploaded photos/documents.
        for idx, player in enumerate(team.get("players", []), start=1):
            player_folder = clean_name(f"{idx:02d} - {player.get('name', 'player')}", f"player_{idx:02d}")
            for kind in FILE_FIELDS:
                rel = (player.get("files", {}) or {}).get(kind)
                if not rel:
                    continue
                src_path = DATA_DIR / rel
                if not src_path.exists() or not src_path.is_file():
                    continue
                ext = src_path.suffix or ".jpg"
                arc_name = f"{team_folder}/photos/{player_folder}/{file_labels.get(kind, kind)}{ext}"
                zf.write(src_path, arc_name)

    zip_buffer.seek(0)
    filename = f"warzone_team_{team_id[:8]}.zip"
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/api/registrations/{team_id}")
def get_registration(team_id: str, request: Request):
    require_admin(request)
    data = load_data()
    for team in data.get("teams", []):
        if team.get("id") == team_id:
            item = public_team(team, request=request)
            whatsapp = group_for_team(team_id)
            item["whatsapp_group"] = {
                "slot": whatsapp.get("slot"),
                "name": whatsapp.get("name"),
                "link": whatsapp.get("link"),
            } if whatsapp else None
            return item
    raise HTTPException(status_code=404, detail="الفريق غير موجود.")


@router.put("/api/registrations/{team_id}")
async def update_registration(team_id: str, request: Request):
    require_admin(request)
    data = load_data()
    for idx, old_team in enumerate(data.get("teams", [])):
        if old_team.get("id") == team_id:
            updated = await build_team_from_form(request, existing=old_team)
            ensure_team_name_unique(data, updated["team_name"], exclude_team_id=team_id)

            # Check national IDs across other teams.
            other_ids = {p.get("national_id") for t in data.get("teams", []) if t.get("id") != team_id for p in t.get("players", [])}
            for p in updated["players"]:
                if p.get("national_id") in other_ids:
                    raise HTTPException(status_code=409, detail=f"الرقم القومي {p.get('national_id')} مسجل في فريق آخر.")

            data["teams"][idx] = updated
            save_data(data)
            return {"status": "success", "team_id": team_id}
    raise HTTPException(status_code=404, detail="الفريق غير موجود.")


@router.delete("/api/registrations/{team_id}")
def delete_registration(team_id: str, request: Request):
    require_admin(request)
    data = load_data()
    before = len(data.get("teams", []))
    data["teams"] = [t for t in data.get("teams", []) if t.get("id") != team_id]
    if len(data["teams"]) == before:
        raise HTTPException(status_code=404, detail="الفريق غير موجود.")
    delete_team_files(team_id)
    release_whatsapp_group(team_id)
    save_data(data)
    return {"status": "success"}


@router.get("/api/registration-file/{team_id}/{player_id}/{kind}")
def get_registration_file(team_id: str, player_id: str, kind: str, request: Request):
    require_admin(request)
    if kind not in FILE_FIELDS:
        raise HTTPException(status_code=404, detail="نوع الملف غير صحيح.")
    data = load_data()
    for team in data.get("teams", []):
        if team.get("id") == team_id:
            for player in team.get("players", []):
                if player.get("id") == player_id:
                    rel = player.get("files", {}).get(kind)
                    if not rel:
                        raise HTTPException(status_code=404, detail="الملف غير موجود.")
                    path = DATA_DIR / rel
                    if not path.exists():
                        raise HTTPException(status_code=404, detail="الملف غير موجود على السيرفر.")
                    return FileResponse(path)
    raise HTTPException(status_code=404, detail="الملف غير موجود.")
